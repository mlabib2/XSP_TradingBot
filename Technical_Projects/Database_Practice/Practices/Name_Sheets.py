import os
import re
from openpyxl import load_workbook

ROOT = "/Users/mahirlabib/Desktop/Darwin_Buffet_Stock_Pick_Project-2/Financials_Analyzed/HK_Stock"

KEY_SHEETS = {
    "balance", "balance_sheet",
    "income", "income_statement",
    "cash_flow", "cash flow", "cashflow",
    "growth", "growth_ability",
    "dupont", "dupont_analysis",
    "profitability", "earning_quality", "profitability_and_earning"
}

def normalize(s: str) -> str:
    # lower, strip non-alnum
    return re.sub(r'[^a-z0-9]', '', s.lower())

def sheet_signals(sheetnames):
    """How 'final' a workbook looks: count key-word hits in sheet names."""
    score = 0
    for sn in sheetnames:
        snn = normalize(sn)
        for k in KEY_SHEETS:
            if normalize(k) in snn:
                score += 1
                break
    return score

def pick_final_xlsx(folder_path: str) -> str | None:
    """Pick the best 'final' workbook in a folder."""
    folder = os.path.basename(folder_path)
    folder_norm = normalize(folder)

    xlsxs = [f for f in os.listdir(folder_path) if f.lower().endswith(".xlsx")]
    if not xlsxs:
        return None

    # Build candidates with scores
    candidates = []
    for fname in xlsxs:
        stem = os.path.splitext(fname)[0]
        stem_norm = normalize(stem)

        # Name score: exact > contains/contained > none
        if stem_norm == folder_norm:
            name_score = 3
        elif stem_norm in folder_norm or folder_norm in stem_norm:
            name_score = 2
        else:
            name_score = 0

        # Size for tie-break
        fpath = os.path.join(folder_path, fname)
        try:
            fsize = os.path.getsize(fpath)
        except:
            fsize = 0

        # Read sheetnames (safely)
        try:
            wb = load_workbook(fpath, read_only=True, data_only=True)
            sheets = wb.sheetnames
            wb.close()
        except Exception:
            sheets = []

        candidates.append({
            "path": fpath,
            "name_score": name_score,
            "sheet_signal": sheet_signals(sheets),
            "num_sheets": len(sheets),
            "size": fsize,
            "sheets": sheets
        })

    # Sort by: name_score desc, sheet_signal desc, num_sheets desc, size desc
    candidates.sort(key=lambda c: (c["name_score"], c["sheet_signal"], c["num_sheets"], c["size"]), reverse=True)

    best = candidates[0]
    # Require at least a weak name match to avoid picking “Balance_Sheet.xlsx” as final
    return best["path"] if best["name_score"] >= 2 else None

printed = []
skipped = []

# Only iterate immediate company subfolders
for entry in sorted(os.listdir(ROOT)):
    company_dir = os.path.join(ROOT, entry)
    if not os.path.isdir(company_dir):
        continue

    final_path = pick_final_xlsx(company_dir)
    if final_path:
        try:
            wb = load_workbook(final_path, read_only=True, data_only=True)
            print(f"\nFile: {final_path}")
            for sn in wb.sheetnames:
                print(f"  - {sn}")
            wb.close()
            printed.append(entry)
        except Exception as e:
            print(f"❌ Could not read {final_path}: {e}")
            skipped.append(entry)
    else:
        skipped.append(entry)

# Debug summary (helps find any folder we didn’t print)
if skipped:
    print("\n--- Skipped folders (no final workbook detected) ---")
    for name in skipped:
        print(f"- {name}")