import os
import sys
import shutil
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook

# === change if your root moves ===
ROOT = Path("/Users/mahirlabib/Desktop/Darwin_Buffet_Stock_Pick_Project-2/Financials_Analyzed/HK_Stock")
SHEET_NAME = "Stock Pick Metrics"

# EXACT content you provided (labels, formulas, numbers)
CELLS = [
    ("A33", "Score"),

    ("A35", "5 year Op income growth (15 points)"),
    ("B35", "=$A41*MIN((B13-B41)/(C41-B41),1)"),

    ("A36", "5 year net asset growth (20 points)"),
    ("B36", "=$A42*MIN((B17-B42)/(C42-B42),1)"),

    ("A37", "5 year avg ROCE (30 points)"),
    ("B37", "=$A43*MIN((AVERAGE(B20:F20)-B43)/(C43-B43),1)"),

    ("A38", "5 year avg EBIT margin (15 points)"),
    ("B38", "=$A44*MIN((AVERAGE(B25:F25)-B44)/(C44-B44),1)"),

    # keep the user's spelling "avergae"
    ("A39", "5 year avergae gearing (20 points)"),
    ("B39", "=$A45*MIN((B45-AVERAGE(B29:F29))/(B45-C45),1)"),

    # constants table A41:C45
    ("A41", 15), ("B41", 2),  ("C41", 30),
    ("A42", 20), ("B42", 2),  ("C42", 20),
    ("A43", 30), ("B43", 6),  ("C43", 20),
    ("A44", 15), ("B44", 10), ("C44", 30),
    ("A45", 20), ("B45", 50), ("C45", 20),
]

def is_tmp_excel(name: str) -> bool:
    return name.startswith("~$") or name.startswith("._")

def backup_file(path: Path, backups_dir: Path) -> Path:
    backups_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    dest = backups_dir / f"{path.stem}.{stamp}{path.suffix}"
    shutil.copy2(path, dest)
    return dest

def apply_cells(ws):
    for addr, value in CELLS:
        ws[addr].value = value

def update_workbook(xlsx_path: Path) -> tuple[bool, str]:
    try:
        wb = load_workbook(filename=xlsx_path, data_only=False)
    except Exception as e:
        return False, f"❌ open failed: {e}"

    if SHEET_NAME not in wb.sheetnames:
        return False, "— sheet missing, skipped"

    ws = wb[SHEET_NAME]
    apply_cells(ws)

    try:
        wb.save(xlsx_path)
        return True, "✅ updated"
    except Exception as e:
        return False, f"❌ save failed: {e}"

def main(root: Path):
    backups = root / "_backups_stock_pick_metrics"
    updated, failed = 0, 0
    details = []

    for xlsx in root.rglob("*.xlsx"):
        if is_tmp_excel(xlsx.name):
            continue

        # quick pre-check: only touch files that actually have the target sheet
        try:
            wb = load_workbook(filename=xlsx, read_only=True)
            if SHEET_NAME not in wb.sheetnames:
                wb.close()
                continue
            wb.close()
        except Exception:
            # if read-only check fails, proceed and let update_workbook report
            pass

        try:
            backup = backup_file(xlsx, backups)
            ok, msg = update_workbook(xlsx)
            if ok:
                updated += 1
            else:
                failed += 1
            details.append(f"{xlsx} -> {msg} (backup: {backup.name})")
        except Exception as e:
            failed += 1
            details.append(f"{xlsx} -> ❌ unexpected error: {e}")

    print("\n=== Stock Pick Metrics bulk apply ===")
    print(f"Root: {root}")
    print(f"Backups: {backups}")
    print(f"Updated: {updated}   Failed: {failed}")
    print("\nDetails:")
    for line in details:
        print(line)

if __name__ == "__main__":
    custom_root = Path(sys.argv[1]).expanduser() if len(sys.argv) > 1 else ROOT
    main(custom_root)