import os
from openpyxl import load_workbook, Workbook
from copy import copy

root_dir = "/Users/mahirlabib/Desktop/Sample_Merge"

for folder in os.listdir(root_dir):
    folder_path = os.path.join(root_dir, folder)
    if not os.path.isdir(folder_path):
        continue

    # Create combined workbook
    output_file = os.path.join(folder_path, f"{folder}.xlsx")
    output_wb = Workbook()
    # Remove default sheet if present
    default_sheet = output_wb.active
    output_wb.remove(default_sheet)

    for file in os.listdir(folder_path):
        if not file.lower().endswith('.xlsx'):
            continue
        if file == f"{folder}.xlsx":
            continue

        file_path = os.path.join(folder_path, file)
        try:
            src_wb = load_workbook(file_path, data_only=False)
        except Exception as e:
            print(f"Error loading {file_path}: {e}")
            continue

        for src_sheet in src_wb.worksheets:
            # Create new sheet with truncated name
            sheet_name = os.path.splitext(file)[0][:31]
            new_sheet = output_wb.create_sheet(title=sheet_name)

            # Copy column widths
            for col_letter, dim in src_sheet.column_dimensions.items():
                new_sheet.column_dimensions[col_letter].width = dim.width
            # Copy row heights
            for row_idx, dim in src_sheet.row_dimensions.items():
                if dim.height:
                    new_sheet.row_dimensions[row_idx].height = dim.height

            # Copy cell values and formatting
            for row in src_sheet.iter_rows():
                for cell in row:
                    new_cell = new_sheet.cell(row=cell.row, column=cell.col_idx, value=cell.value)
                    if cell.has_style:
                        new_cell.font = copy(cell.font)
                        new_cell.border = copy(cell.border)
                        new_cell.fill = copy(cell.fill)
                        new_cell.number_format = cell.number_format
                        new_cell.protection = copy(cell.protection)
                        new_cell.alignment = copy(cell.alignment)

            # Copy merged cell ranges
            for merged_range in src_sheet.merged_cells.ranges:
                new_sheet.merge_cells(str(merged_range))

    # Save the combined workbook for this folder
    output_wb.save(output_file)
